import openpyxl
import numpy as np
import statistics
import matplotlib.pyplot as plt
from pathlib import Path

book = openpyxl.open(str(Path(__file__).parent / 'Hepatitis_dataset.xlsx'))

sheet = book.active

ALT_blood_donor = []
for row in sheet.iter_rows(min_row=2, max_row=534, min_col=7, max_col=7):
    for cell in row:
        ALT_blood_donor.append(cell.value)

AST_blood_donor = []
for row in sheet.iter_rows(min_row=2, max_row=534, min_col=8, max_col=8):
    for cell in row:
        AST_blood_donor.append(cell.value)

BIL_blood_donor = []
for row in sheet.iter_rows(min_row=2, max_row=534, min_col=9, max_col=9):
    for cell in row:
        BIL_blood_donor.append(cell.value)

ALT_suspect_donor = []
for row in sheet.iter_rows(min_row=535, max_row=541, min_col=7, max_col=7):
    for cell in row:
        ALT_suspect_donor.append(cell.value)

AST_suspect_donor = []
for row in sheet.iter_rows(min_row=535, max_row=541, min_col=8, max_col=8):
    for cell in row:
        AST_suspect_donor.append(cell.value)

BIL_suspect_donor = []
for row in sheet.iter_rows(min_row=535, max_row=541, min_col=9, max_col=9):
    for cell in row:
        BIL_suspect_donor.append(cell.value)

ALT_hepatitis = []
for row in sheet.iter_rows(min_row=541, max_row=565, min_col=7, max_col=7):
    for cell in row:
        ALT_hepatitis.append(cell.value)

AST_hepatitis = []
for row in sheet.iter_rows(min_row=541, max_row=565, min_col=8, max_col=8):
    for cell in row:
        AST_hepatitis.append(cell.value)

BIL_hepatitis = []
for row in sheet.iter_rows(min_row=541, max_row=565, min_col=9, max_col=9):
    for cell in row:
        BIL_hepatitis.append(cell.value)

ALT_fibrosis = []
for row in sheet.iter_rows(min_row=565, max_row=586, min_col=7, max_col=7):
    for cell in row:
        ALT_fibrosis.append(cell.value)

AST_fibrosis = []
for row in sheet.iter_rows(min_row=565, max_row=586, min_col=8, max_col=8):
    for cell in row:
        AST_fibrosis.append(cell.value)

BIL_fibrosis = []
for row in sheet.iter_rows(min_row=565, max_row=586, min_col=9, max_col=9):
    for cell in row:
        BIL_fibrosis.append(cell.value)

ALT_cirrhosis = []
for row in sheet.iter_rows(min_row=586, max_row=616, min_col=7, max_col=7):
    for cell in row:
        ALT_cirrhosis.append(cell.value)

AST_cirrhosis = []
for row in sheet.iter_rows(min_row=586, max_row=616, min_col=8, max_col=8):
    for cell in row:
        AST_cirrhosis.append(cell.value)

BIL_cirrhosis = []
for row in sheet.iter_rows(min_row=586, max_row=616, min_col=9, max_col=9):
    for cell in row:
        BIL_cirrhosis.append(cell.value)

mean_ALT_blood_donor = np.mean(ALT_blood_donor)
print(f'Mean ALT of Blood donors is {mean_ALT_blood_donor}')

mean_AST_blood_donor = np.mean(AST_blood_donor)
print(f'Mean AST of Blood donors is {mean_ALT_blood_donor}')

mean_BIL_blood_donor = np.mean(BIL_blood_donor)
print(f'Mean BIL of Blood donors is {mean_BIL_blood_donor}')

print()

mean_ALT_suspect_donor = np.mean(ALT_suspect_donor)
print(f'Mean ALT of Suspect_donors is {mean_ALT_suspect_donor}')

mean_AST_suspect_donor = np.mean(AST_suspect_donor)
print(f'Mean AST of Suspect_donors is {mean_AST_suspect_donor}')

mean_BIL_suspect_donor = np.mean(ALT_suspect_donor)
print(f'Mean BIL of Suspect_donors is {mean_BIL_suspect_donor}')

print()

mean_ALT_hepatitis = np.mean(ALT_hepatitis)
print(f'Mean ALT of patients with Hepatitis is {mean_ALT_hepatitis}')

mean_AST_hepatitis = np.mean(AST_hepatitis)
print(f'Mean AST of patients with Hepatitis is {mean_AST_hepatitis}')

mean_BIL_hepatitis = np.mean(BIL_hepatitis)
print(f'Mean BIL of patients with Hepatitis is {mean_BIL_hepatitis}')

print()

mean_ALT_fibrosis = np.mean(ALT_fibrosis)
print(f'Mean ALT of patients with Fibrosis is {mean_ALT_fibrosis}')

mean_AST_fibrosis = np.mean(AST_fibrosis)
print(f'Mean AST of patients with Fibrosis is {mean_AST_fibrosis}')

mean_BIL_fibrosis = np.mean(BIL_fibrosis)
print(f'Mean BIL of patients with Fibrosis is {mean_BIL_fibrosis}')

print()

mean_ALT_cirrhosis = np.mean(ALT_cirrhosis)
print(f'Mean ALT of patients with Cirrhosis is {mean_ALT_cirrhosis}')

mean_AST_cirrhosis = np.mean(AST_cirrhosis)
print(f'Mean AST of patients with Cirrhosis is {mean_AST_cirrhosis}')

mean_BIL_cirrhosis = np.mean(BIL_cirrhosis)
print(f'Mean BIL of patients with Cirrhosis is {mean_BIL_cirrhosis }')

print()

dataset_ALT_blood_donor = np.array([ALT_blood_donor])
std_ALT_blood_donor = dataset_ALT_blood_donor.std(ddof=0)
print(f'Standard deviation of values ALT of Blood donors is {std_ALT_blood_donor}')

dataset_ALT_suspect_donor = np.array([ALT_suspect_donor])
std_ALT_suspect_donor = dataset_ALT_suspect_donor.std(ddof=1)
print(f'Standard deviation of values ALT of Suspect donors is {std_ALT_suspect_donor}')

dataset_ALT_hepatitis = np.array([ALT_hepatitis])
std_ALT_hepatitis = dataset_ALT_hepatitis.std(ddof=1)
print(f'Standard deviation of values ALT of patients with Hepatitis is {std_ALT_hepatitis}')

dataset_ALT_fibrosis = np.array([ALT_fibrosis])
std_ALT_fibrosis = dataset_ALT_fibrosis.std(ddof=1)
print(f'Standard deviation of values ALT of patients with Fibrosis is {std_ALT_fibrosis}')

dataset_ALT_cirrhosis = np.array([ALT_cirrhosis])
std_ALT_cirrhosis = dataset_ALT_cirrhosis.std(ddof=1)
print(f'Standard deviation of values ALT of patients with Cirrhosis is {std_ALT_cirrhosis}')

print()

variance_ALT_blood_donor = std_ALT_blood_donor ** 2
print(f'Variance of values ALT of Blood donors is {variance_ALT_blood_donor}')

variance_ALT_suspect_donor = std_ALT_suspect_donor ** 2
print(f'Variance of values ALT of Suspect_donors is {variance_ALT_suspect_donor}')

variance_ALT_hepatitis = std_ALT_hepatitis ** 2
print(f'Variance of values ALT of patients with Hepatitis is {variance_ALT_hepatitis}')

variance_ALT_fibrosis = std_ALT_fibrosis ** 2
print(f'Variance of values ALT of patients with Fibrosis is {variance_ALT_fibrosis}')

variance_ALT_cirrhosis = std_ALT_cirrhosis ** 2
print(f'Variance of values ALT of patients with Cirrhosis is {variance_ALT_cirrhosis}')

print()

median_ALT_blood_donor = statistics.median(ALT_blood_donor)
print(f'Median of values ALT of Blood donors is {median_ALT_blood_donor}')

median_ALT_suspect_donor = statistics.median(ALT_suspect_donor)
print(f'Median of values ALT of Suspect donors is {median_ALT_suspect_donor}')

median_ALT_hepatitis = statistics.median(ALT_hepatitis)
print(f'Median of values ALT of patients with Hepatitis is {median_ALT_hepatitis}')

median_ALT_fibrosis = statistics.median(ALT_fibrosis)
print(f'Median of values ALT of patients with Fibrosis is {median_ALT_fibrosis}')

median_ALT_cirrhosis = statistics.median(ALT_cirrhosis)
print(f'Median of values ALT of patients with Cirrhosis is {median_ALT_cirrhosis}')

print()

ALT_blood_donor.sort()

n_ALT_blood_donor = len(ALT_blood_donor)
k = 75

test_number = n_ALT_blood_donor * k % 100

if test_number > 0:
    j_ALT_blood_donor = n_ALT_blood_donor * k // 100
else:
    j_ALT_blood_donor = int((n_ALT_blood_donor * k / 100) + (n_ALT_blood_donor * k / (100 + 1))) // 2 + 1

quart3_ALT_blood_donor = ALT_blood_donor[j_ALT_blood_donor - 1]

print(f'The third quartile of values ALT of Blood donors is {quart3_ALT_blood_donor}')

ALT_suspect_donor.sort()

n_ALT_suspect_donor = len(ALT_suspect_donor)
k = 75

test_number = n_ALT_suspect_donor * k % 100

if test_number > 0:
    j_ALT_suspect_donor = n_ALT_suspect_donor * k // 100
else:
    j_ALT_suspect_donor = int((n_ALT_suspect_donor * k / 100) + (n_ALT_suspect_donor * k / (100 + 1))) // 2 + 1

quart3_ALT_suspect_donor = ALT_suspect_donor[j_ALT_suspect_donor - 1]

print(f'The third quartile of values ALT of Suspect donors is {quart3_ALT_suspect_donor}')

ALT_hepatitis.sort()

n_ALT_hepatitis = len(ALT_hepatitis)
k = 75

test_number = n_ALT_hepatitis * k % 100

if test_number > 0:
    j_ALT_hepatitis = n_ALT_hepatitis * k // 100
else:
    j_ALT_hepatitis = int((n_ALT_hepatitis * k / 100) + (n_ALT_hepatitis * k / (100 + 1))) // 2 + 1

quart3_ALT_hepatitis = ALT_hepatitis[j_ALT_hepatitis - 1]

print(f'The third quartile of values ALT of patients with Hepatitis is {quart3_ALT_hepatitis}')

ALT_fibrosis.sort()

n_ALT_fibrosis = len(ALT_fibrosis)
k = 75

test_number = n_ALT_fibrosis * k % 100

if test_number > 0:
    j_ALT_fibrosis = n_ALT_fibrosis * k // 100
else:
    j_ALT_fibrosis = int((n_ALT_fibrosis * k / 100) + (n_ALT_fibrosis * k / (100 + 1))) // 2 + 1

quart3_ALT_fibrosis = ALT_fibrosis[j_ALT_fibrosis - 1]

print(f'The third quartile of values ALT of patients with Fibrosis is {quart3_ALT_fibrosis}')

ALT_cirrhosis.sort()

n_ALT_cirrhosis = len(ALT_cirrhosis)

test_number = n_ALT_cirrhosis * k % 100

if test_number > 0:
    j_ALT_cirrhosis = n_ALT_cirrhosis * k // 100
else:
    j_ALT_cirrhosis = int((n_ALT_cirrhosis * k / 100) + (n_ALT_cirrhosis * k / (100 + 1))) // 2 + 1

quart3_ALT_cirrhosis = ALT_cirrhosis[j_ALT_cirrhosis - 1]

print(f'The third quartile of values ALT of patients with Cirrhosis is {quart3_ALT_cirrhosis}')

print('')

plt.xlabel('ALT, U/l')
plt.ylabel('Number of patients')
plt.title('ALT')
plt.hist(ALT_blood_donor, label='Healthy blood donors')
plt.hist(ALT_suspect_donor, label='Suspect donors')
plt.hist(ALT_hepatitis, label='Patients with Hepatitis')
plt.hist(ALT_fibrosis, label='Patients with Fibrosis')
plt.hist(ALT_cirrhosis, label='Patients with Cirrhosis')
plt.legend(fontsize=14)
plt.tight_layout()
plt.show()

plt.xlabel('AST, U/l')
plt.ylabel('Number of patients')
plt.title('AST')
plt.hist(AST_blood_donor, label='Healthy blood donors')
plt.hist(AST_suspect_donor, label='Suspect donors')
plt.hist(AST_hepatitis, label='Patients with Hepatitis')
plt.hist(AST_fibrosis, label='Patients with Fibrosis')
plt.hist(AST_cirrhosis, label='Patients with Cirrhosis')
plt.legend(fontsize=14)
plt.tight_layout()
plt.show()

plt.xlabel('BIL, U/l')
plt.ylabel('Number of patients')
plt.title('BIL')
plt.hist(BIL_blood_donor, label='Healthy blood donors')
plt.hist(BIL_suspect_donor, label='Suspect donors')
plt.hist(BIL_hepatitis, label='Patients with Hepatitis')
plt.hist(BIL_fibrosis, label='Patients with Fibrosis')
plt.hist(BIL_cirrhosis, label='Patients with Cirrhosis')
plt.legend(fontsize=14)
plt.tight_layout()
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('ALT, U/l')
plt.title('Mean ALT of all donors and patients')
values = [mean_ALT_blood_donor, mean_ALT_suspect_donor, mean_ALT_hepatitis, mean_ALT_fibrosis, mean_ALT_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('AST, U/l')
plt.title('Mean AST of all donors and patients')
values = [mean_AST_blood_donor, mean_AST_suspect_donor, mean_AST_hepatitis, mean_AST_fibrosis, mean_AST_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('BIL, U/l')
plt.title('Mean BIL of all donors and patients')
values = [mean_BIL_blood_donor, mean_BIL_suspect_donor, mean_BIL_hepatitis, mean_BIL_fibrosis, mean_BIL_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('Standard deviation')
plt.title('Standard deviation of values ALT')
values = [std_ALT_blood_donor, std_ALT_suspect_donor, std_ALT_hepatitis, std_ALT_fibrosis, std_ALT_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('Variance of values')
plt.title('Variance of values ALT')
values = [variance_ALT_blood_donor, variance_ALT_suspect_donor, variance_ALT_hepatitis, variance_ALT_fibrosis, variance_ALT_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('Median of values')
plt.title('Median of values ALT')
values = [median_ALT_blood_donor, median_ALT_suspect_donor, median_ALT_hepatitis, median_ALT_fibrosis, median_ALT_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()

index = [0, 1, 2, 3, 4]
plt.xlabel('Group of patients')
plt.ylabel('The third quartiles')
plt.title('The third quartile of values ALT')
values = [quart3_ALT_blood_donor, quart3_ALT_suspect_donor, quart3_ALT_hepatitis, quart3_ALT_fibrosis, quart3_ALT_cirrhosis]
plt.bar(index, values)
plt.xticks(index, ['HBD', 'SD', 'PH', 'PF', 'PC'])
plt.show()
